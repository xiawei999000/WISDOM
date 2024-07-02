'''
step 1
extract the 2D bounding box (32X32) containing lymph node using zero padding
and save in nii.gz file for each patient
date: 2021-01-21
author: xiav
'''
import numpy as np
import SimpleITK as sitk
import os
import openpyxl
import pickle
import matplotlib.pyplot as plt
import pylab
import LN_functions

def connected_domain_2D(image):
    cca = sitk.ConnectedComponentImageFilter()
    cca.SetFullyConnected(True)
    _input = sitk.GetImageFromArray(image.astype(np.uint8))
    output_ex = cca.Execute(_input)
    output = sitk.GetArrayFromImage(output_ex)

    stats = sitk.LabelShapeStatisticsImageFilter()
    stats.Execute(output_ex)
    num_label = cca.GetObjectCount()
    num_list = [i for i in range(1, num_label+1)]

    # for one_label in num_list:
    #     x, y, w, h = stats.GetBoundingBox(one_label)
    #     one_mask = (output[y: y + h, x: x + w] == one_label)
    #     output[y: y + h, x: x + w] *= one_mask
    return [num_list, stats, output]


if __name__ == '__main__':
    modality = 'T2'
    format = '.nii.gz'
    center_name = 'YUNNAN' # SHANXI FUDAN YUNNAN
    path_ori = 'R:\\Colorectal_cancer\\rectal_cancer_LN_diagnosis\\' + center_name + '\\normalize_all\\' + modality
    path_mask = 'R:\\Colorectal_cancer\\rectal_cancer_LN_diagnosis\\' + center_name + '\\correct_all\\mask'
    path_patchs = './processed_data/' + center_name + '/patches_2d_all/' + modality
    info_path = './processed_data/' + center_name + '/info-all-patches/LN_patch_img_info.xlsx'

    if not (os.path.exists(path_patchs)):
        os.mkdir(path_patchs)

    img_folder_list = os.listdir(path_ori)

    # store all the LN info for statistic
    y_list = []
    x_list = []

    # h_list = []
    # w_list = []

    # INFO
    # data sheet
    info = openpyxl.Workbook()  # 创建一个工作表
    ws = info.active  # ws操作sheet页
    all_sheet = info.create_sheet('1', 0)

    all_sheet.cell(1, 1, 'ID')
    all_sheet.cell(1, 2, 'P_name')
    all_sheet.cell(1, 3, 'P_LN_index')
    all_sheet.cell(1, 4, 'LN_patches_num')
    all_sheet.cell(1, 5, 'x')
    all_sheet.cell(1, 6, 'y')
    all_sheet.cell(1, 7, 'slice_ind')

    index = 1
    total_LN_nums = 0

    # patient level
    for vol_name in img_folder_list:
        # vol_name = 'T2_203zjc_2015_hm.nii.gz'
        p_img_path = path_ori + '\\' + vol_name
        p_img = sitk.ReadImage(p_img_path)
        p_img_array = sitk.GetArrayFromImage(p_img)

        p_mask_path = path_mask + '\\' + (vol_name.replace('_hm', '')).replace(modality, 'mask')
        p_mask = sitk.ReadImage(p_mask_path)
        p_mask_array = sitk.GetArrayFromImage(p_mask)
        volume_shape = p_mask_array.shape

        slice_num = volume_shape[0]
        LN_nums = 0

        P_LN_patches_img = []
        P_LN_patches_mask = []
        P_LN_patches_area = []

        P_LN_index = 0

        P_name = (vol_name.replace('_hm.nii.gz', '')).replace(modality + '_', '')

        P_LN_patches_path_i = path_patchs + '\\' + vol_name

        # number of LN patches
        LN_patches_num = 0
        for slice_ind in range(0, slice_num):
            p_mask_array_slice = p_mask_array[slice_ind]
            if np.max(p_mask_array_slice) == 1:
                [connReg_list, stats, mask_array_con] = connected_domain_2D(p_mask_array_slice)
                LN_patches_num = LN_patches_num + len(connReg_list)

        for slice_ind in range(0, slice_num):
            p_mask_array_slice = p_mask_array[slice_ind]
            if np.max(p_mask_array_slice) == 1:
                # identify each LN by connected_domain in slice
                [connReg_list, stats, mask_array_con] = connected_domain_2D(p_mask_array_slice)

                LN_nums = len(connReg_list) + LN_nums
                # extend the boarder of the mask to contain the tumor border information
                spacing = 4

                # size for each LN
                for one_label in connReg_list:
                    x, y, w, h = stats.GetBoundingBox(one_label)

                    # h_list.append(h)
                    # w_list.append(w)

                    y_list.append(y)
                    x_list.append(x)

                    y1 = y - spacing
                    y2 = y + h + spacing

                    x1 = x - spacing
                    x2 = x + w + spacing

                    if y1 < 0:
                        y1 = 0
                    if x1 < 0:
                        x1 = 0

                    img_size = p_mask_array.shape
                    img_width = img_size[1]
                    img_height = img_size[2]

                    if y2 > img_height:
                        y2 = img_height
                    if x2 > img_width:
                        x2 = img_width

                    p_slice_bounding = p_img_array[slice_ind, y1: y2, x1: x2]

                    # p_mask_bounding = p_mask_array[slice_ind, y1: y2, x1: x2]
                    # plt.subplot(221)
                    # plt.imshow(p_slice_bounding, cmap='gray')
                    # plt.subplot(222)
                    # plt.imshow(p_mask_bounding, cmap='gray')
                    # plt.subplot(223)
                    # plt.imshow(adc_slice_bounding, cmap='gray')
                    # plt.subplot(224)
                    # plt.imshow(mask_slice_bounding, cmap='gray')
                    # plt.show()

                    P_LN_patches_img.append(p_slice_bounding)
                    # P_LN_patches_mask.append(p_mask_bounding)
                    # P_LN_patches_area.append([h, w])

                    # store the information of each LN node patch
                    index = index + 1
                    all_sheet.cell(index, 1, index-1)
                    all_sheet.cell(index, 2, P_name)
                    all_sheet.cell(index, 3, P_LN_index)
                    all_sheet.cell(index, 4, LN_patches_num)
                    all_sheet.cell(index, 5, x)
                    all_sheet.cell(index, 6, y)
                    all_sheet.cell(index, 7, slice_ind)
                    P_LN_index = P_LN_index + 1

        P_LN_patches_img_padding = LN_functions.get_pLevel_LN_patches_Resize(P_LN_patches_img)
        P_LN_patches_img_Image = sitk.GetImageFromArray(P_LN_patches_img_padding)

        sitk.WriteImage(P_LN_patches_img_Image, P_LN_patches_path_i)

        total_LN_nums = total_LN_nums + LN_nums
        print('pid with ', P_name, ' has ', LN_nums, ' LNs.')
    print('total number of LNs is :', str(total_LN_nums))
    info.save(filename=info_path)





