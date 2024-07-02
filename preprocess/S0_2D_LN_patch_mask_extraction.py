'''
step 0
extract the 2D bounding box (32X32) containing lymph node mask using zero padding
and save in nii.gz file for each patient
date: 2021-09-15
author: xiav
'''
import numpy as np
import SimpleITK as sitk
import os
import openpyxl
import pickle
import matplotlib.pyplot as plt
import pylab
import LN_functions

def connected_domain_2D(image):
    cca = sitk.ConnectedComponentImageFilter()
    cca.SetFullyConnected(True)
    _input = sitk.GetImageFromArray(image.astype(np.uint8))
    output_ex = cca.Execute(_input)
    output = sitk.GetArrayFromImage(output_ex)

    stats = sitk.LabelShapeStatisticsImageFilter()
    stats.Execute(output_ex)
    num_label = cca.GetObjectCount()
    num_list = [i for i in range(1, num_label+1)]

    # for one_label in num_list:
    #     x, y, w, h = stats.GetBoundingBox(one_label)
    #     one_mask = (output[y: y + h, x: x + w] == one_label)
    #     output[y: y + h, x: x + w] *= one_mask
    return [num_list, stats, output]


if __name__ == '__main__':
    center_name = 'YUNNAN' # SHANXI FUDAN YUNNAN
    format = '.nii.gz'
    path_mask = 'R:\\Colorectal_cancer\\rectal_cancer_LN_diagnosis\\' + center_name + '\\correct_all\\mask'
    path_patches = './processed_data/' + center_name + '/patches_2d_all/mask'
    info_path = './processed_data/' + center_name + '/info-all-patches/LN_patch_mask_info.xlsx'

    if not (os.path.exists(path_patches)):
        os.mkdir(path_patches)

    img_folder_list = os.listdir(path_mask)

    # INFO
    # data sheet
    info = openpyxl.Workbook()  # 创建一个工作表
    ws = info.active  # ws操作sheet页
    all_sheet = info.create_sheet('1', 0)

    all_sheet.cell(1, 1, 'ID')
    all_sheet.cell(1, 2, 'P_name')
    all_sheet.cell(1, 3, 'P_LN_num')
    all_sheet.cell(1, 4, 'x')
    all_sheet.cell(1, 5, 'y')
    all_sheet.cell(1, 6, 'slice_ind')

    index = 1
    total_LN_nums = 0

    # patient level
    for vol_name in img_folder_list:
        p_mask_path = path_mask + '\\' + vol_name
        p_mask = sitk.ReadImage(p_mask_path)
        p_mask_array = sitk.GetArrayFromImage(p_mask)
        vol_spacing = p_mask.GetSpacing()
        volume_shape = p_mask_array.shape

        slice_num = volume_shape[0]
        LN_nums = 0

        P_LN_patches_mask = []
        P_LN_patches_area = []

        P_LN_num = 0

        P_name = (vol_name.replace('.nii.gz', '')).replace('mask_', '')

        P_LN_patches_path_i = path_patches + '\\' + vol_name

        # if not (os.path.exists(P_LN_patches_path_i)):
        # slice level
        for slice_ind in range(0, slice_num):
            p_mask_array_slice = p_mask_array[slice_ind]

            if np.max(p_mask_array_slice) == 1:
                # identify each LN by connected_domain in slice
                [connReg_list, stats, mask_array_con] = connected_domain_2D(p_mask_array_slice)

                LN_nums = len(connReg_list) + LN_nums
                # extend the boarder of the mask to contain the tumor border information
                spacing = 4 # for saving the LN patches

                # size for each LN
                for one_label in connReg_list:
                    x, y, w, h = stats.GetBoundingBox(one_label)

                    y1 = y - spacing
                    y2 = y + h + spacing

                    x1 = x - spacing
                    x2 = x + w + spacing

                    if y1 < 0:
                        y1 = 0
                    if x1 < 0:
                        x1 = 0

                    img_size = p_mask_array.shape
                    img_width = img_size[1]
                    img_height = img_size[2]

                    if y2 > img_height:
                        y2 = img_height
                    if x2 > img_width:
                        x2 = img_width

                    p_mask_bounding = p_mask_array[slice_ind, y1: y2, x1: x2]


                    P_LN_patches_mask.append(p_mask_bounding)

                    # store the information of each LN node patch
                    index = index + 1
                    all_sheet.cell(index, 1, index-1)
                    all_sheet.cell(index, 2, P_name)
                    all_sheet.cell(index, 3, P_LN_num)
                    all_sheet.cell(index, 4, x)
                    all_sheet.cell(index, 5, y)
                    all_sheet.cell(index, 6, slice_ind)
                    P_LN_num = P_LN_num + 1

        P_LN_patches_mask_padding = LN_functions.get_pLevel_LN_patches_Resize(P_LN_patches_mask)
        P_LN_patches_mask_Image = sitk.GetImageFromArray(P_LN_patches_mask_padding)

        sitk.WriteImage(P_LN_patches_mask_Image, P_LN_patches_path_i)

        total_LN_nums = total_LN_nums + LN_nums
        print('pid with ', P_name, ' has ', LN_nums, ' LNs.')

    print('total number of LNs is :', str(total_LN_nums))
    info.save(filename=info_path)
