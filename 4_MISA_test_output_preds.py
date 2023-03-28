# generate the predicion of each patch
import os
import pickle
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.models import load_model
import openpyxl
import numpy as np
import random

my_seed = 666
np.random.seed(my_seed)
random.seed(my_seed)
tf.random.set_seed(my_seed)

os.environ["CUDA_VISIBLE_DEVICES"] = "1"

def patches_feas_zero_padding(patches_feas, patches_num):
    p_num = len(patches_feas)
    patches_feas_padding = np.zeros([p_num, patches_num])
    for ind_p in range(0, p_num):
        P_patches_feas = patches_feas[ind_p]
        patch_num_p = len(P_patches_feas)
        P_patches_feas_temp = np.zeros(patches_num)
        for ind_patch in range(0, patch_num_p):
            LN_feas = P_patches_feas[ind_patch]
            P_patches_feas_temp[ind_patch] = LN_feas
        patches_feas_padding[ind_p, :] = P_patches_feas_temp
    return patches_feas_padding

if __name__ == '__main__':

    center_name = 'CenterI'

    WISDOM_model = load_model('./I_S_A_DiagnosticNetwork.hdf5',
                           compile=False)

    model = keras.Model(inputs=WISDOM_model.input, outputs=WISDOM_model.get_layer('TimeDistributed').output)

    # load data for model test
    patch_num = 80
    patches_preds_img = pickle.load(open(
        './data/' + center_name + '/output_preds/' + 'patches_preds_ImgLevel.bin',
        "rb"))

    patches_short_diameter = pickle.load(open(
        './data/' + center_name + '/patches_2d_bin_std_all/' + 'Patches_short_diameter.bin',
        "rb"))
    patches_short_diameter_padding = patches_feas_zero_padding(patches_short_diameter, patch_num)

    patches_long_diameter = pickle.load(open(
        './data/' + center_name + '/patches_2d_bin_std_all/' + 'Patches_long_diameter.bin',
        "rb"))
    patches_long_diameter_padding = patches_feas_zero_padding(patches_long_diameter, patch_num)

    patches_ratio_diameter = pickle.load(open(
        './data/' + center_name + '/patches_2d_bin_std_all/' + 'Patches_ratio_diameter.bin',
        "rb"))
    patches_ratio_diameter_padding = patches_feas_zero_padding(patches_ratio_diameter, patch_num)

    patches_adc = pickle.load(open(
        './data/' + center_name + '/patches_2d_bin_std_all/' + 'Patches_adc_value.bin',
        "rb"))
    patches_adc_padding = patches_feas_zero_padding(patches_adc, patch_num)

    # combine all data into one
    # feature normalization
    # diameter:10 mm cutoff for LN positive
    # ADC val: max val less than 100, control the feature value within the range of [0,1]
    p_num = patches_preds_img.shape[0]
    feas_num = 5
    all_p_data = np.zeros([p_num, patch_num, feas_num])
    all_p_data[:, :, 0] = patches_preds_img
    all_p_data[:, :, 1] = patches_short_diameter_padding / 10
    all_p_data[:, :, 2] = patches_long_diameter_padding / 10
    all_p_data[:, :, 3] = patches_ratio_diameter_padding
    all_p_data[:, :, 4] = patches_adc_padding / 100
    x_test = np.asarray(all_p_data)
    patches_preds_test = model.predict(x_test)

    # load the info
    all_p_name = pickle.load(open('./data/' + center_name + '/patches_2d_bin_std_all/P_name.bin', "rb"))
    all_LN_meta_labels = pickle.load(open('./data/' + center_name + '/patches_2d_bin_std_all/P_label.bin', "rb"))
    all_p_N_stage = pickle.load(open('./data/' + center_name + '/patches_2d_bin_std_all/P_N_stage.bin', "rb"))
    all_p_N_stage_fine = pickle.load(open('./data/' + center_name + '/patches_2d_bin_std_all/P_N_stage_fine.bin', "rb"))

    # create a new excel table for storing the preds and patient information
    wb = openpyxl.Workbook()
    ws = wb.active
    all_sheet = wb.create_sheet('1', 0)
    # write the table head (name of col)
    all_sheet.cell(1, 1, 'name')
    all_sheet.cell(1, 2, 'patch_id')
    all_sheet.cell(1, 3, 'short_diameter')
    all_sheet.cell(1, 4, 'long_diameter')
    all_sheet.cell(1, 5, 'ratio_diameter')
    all_sheet.cell(1, 6, 'adc_mean_val')
    all_sheet.cell(1, 7, 'patch_pred_img')
    all_sheet.cell(1, 8, 'patch_pred_fusion')
    all_sheet.cell(1, 9, 'label')
    all_sheet.cell(1, 10, 'N_stage')
    all_sheet.cell(1, 11, 'N_stage_fine')
    all_sheet.cell(1, 12, 'set')
    # ---------------------------------------------------------
    set_name = ''
    training_ind = []
    val_ind = []
    if center_name == 'CenterI':
        # data set split
        # arrange the training and validation set
        training_ind = pickle.load(
            open('./data/' + center_name + '/patches_2d_bin_std_all/training_ind_random_666.bin', "rb"))
        val_ind = pickle.load(open('./data/' + center_name + '/patches_2d_bin_std_all/val_ind_random_666.bin', "rb"))
    pid = 1
    for p_index in range(0, len(all_p_name)):
        if center_name == 'CenterI':
            if p_index in training_ind:
                set_name = 'train'
            if p_index in val_ind:
                set_name = 'val'
        else:
            set_name = 'external_test'

        P_patches_feas = patches_adc[p_index]
        patch_num_p = len(P_patches_feas)

        for patch_id in range(0, patch_num_p):
            pid = pid + 1
            all_sheet.cell(pid, 1, all_p_name[p_index])
            all_sheet.cell(pid, 2, patch_id)
            all_sheet.cell(pid, 3, all_p_data[p_index, patch_id, 1])
            all_sheet.cell(pid, 4, all_p_data[p_index, patch_id, 2])
            all_sheet.cell(pid, 5, all_p_data[p_index, patch_id, 3])
            all_sheet.cell(pid, 6, all_p_data[p_index, patch_id, 4])
            all_sheet.cell(pid, 7, all_p_data[p_index, patch_id, 0])
            all_sheet.cell(pid, 8, patches_preds_test[p_index, patch_id, 0])
            all_sheet.cell(pid, 9, all_LN_meta_labels[p_index])
            all_sheet.cell(pid, 10, all_p_N_stage[p_index])
            all_sheet.cell(pid, 11, all_p_N_stage_fine[p_index])
            all_sheet.cell(pid, 12, set_name)
    # -----------------------------------------------
    results_table_save_path = './data/' + center_name + '/output_preds/'
    if not os.path.exists(results_table_save_path):
        os.makedirs(results_table_save_path)
    wb.save(results_table_save_path + center_name + '_' + 'patches_preds.xlsx')