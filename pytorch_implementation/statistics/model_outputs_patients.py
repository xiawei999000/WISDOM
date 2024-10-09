# generate the predicion of each patch in the model stage 2
from __future__ import print_function, division
import torch
import torch.nn as nn
import torch.optim as optim
from torch.optim import lr_scheduler
import numpy as np
import torchvision
from torchvision import transforms as T
from torchvision import datasets, models
import matplotlib.pyplot as plt
import time
import math
import os
import copy
import random
import torch.utils.data as data
import pickle
from sklearn.metrics import roc_auc_score, f1_score, accuracy_score, recall_score, precision_score, roc_curve, confusion_matrix
from _collections import OrderedDict
import openpyxl
from sklearn.model_selection import KFold, StratifiedKFold
# import pandas
# from torch.utils.tensorboard import SummaryWriter
from Self_defined_Resnet_model_1 import ResNetSelf_Combine_ImgFeas_TwoOuts # test for smaller model performance
import monai
from monai.utils import set_determinism


# load data
def load_LN_data(center_name, data_folder_path):
    # load original data
    all_p_name = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_name.bin', "rb"))

    all_patches = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'T2_2D_patches_correct_3channel.bin', "rb"))
    all_patch_LDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_long_diameter.bin', "rb"))
    all_patch_SDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_short_diameter.bin', "rb"))
    all_patch_RDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_ratio_diameter.bin', "rb"))
    all_patch_adc = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_adc_value.bin', "rb"))
    all_LN_meta_labels = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_label.bin', "rb"))
    all_p_N_stages = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_N_stage.bin', "rb"))
    p_set_name = [center_name] * len(all_p_name)
    if center_name == 'FUDAN':
        p_set_name = ['EVC1'] * len(all_p_name)
    if center_name == 'YUNNAN':
        p_set_name = ['EVC2'] * len(all_p_name)

    # # exclude the bad data
    ban_list = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_ban_list.bin', "rb"))
    # ban_list = []
    all_p_ind = list(range(len(all_patches)))
    all_p_include_ind = []
    for ind in all_p_ind:
        if ind not in ban_list:
            all_p_include_ind.append(ind)

    all_p_name = [all_p_name[i] for i in all_p_include_ind]
    all_patches = [all_patches[i] for i in all_p_include_ind]
    all_patch_LDs = [all_patch_LDs[i] for i in all_p_include_ind]
    all_patch_SDs = [all_patch_SDs[i] for i in all_p_include_ind]
    all_patch_RDs = [all_patch_RDs[i] for i in all_p_include_ind]
    all_patch_adc = [all_patch_adc[i] for i in all_p_include_ind]
    all_LN_meta_labels = [all_LN_meta_labels[i] for i in all_p_include_ind]
    all_p_N_stages = [all_p_N_stages[i] for i in all_p_include_ind]

    normalized_all_patch_LDs = [
        [value / 10 for value in sublist]
        for sublist in all_patch_LDs
    ]

    normalized_all_patch_SDs = [
        [value / 10 for value in sublist]
        for sublist in all_patch_SDs
    ]

    normalized_all_patch_adc = [
        [value / 100 for value in sublist]
        for sublist in all_patch_adc
    ]

    # mark the data set
    if center_name == 'SHANXI':
        train_set_ind = pickle.load(
            open(data_folder_path + center_name + '/patches_2d_bin_std_all/training_ind.bin',
                "rb"))
        val_set_ind = pickle.load(
            open(data_folder_path + center_name + '/patches_2d_bin_std_all/val_ind.bin',
                 "rb"))
        for ind in train_set_ind:
            p_set_name[ind] = 'train'
        for ind in val_set_ind:
            p_set_name[ind] = 'val'
        p_set_name = [p_set_name[i] for i in all_p_include_ind]

    return all_p_name, all_patches, normalized_all_patch_LDs, normalized_all_patch_SDs, all_patch_RDs, normalized_all_patch_adc, all_LN_meta_labels, all_p_N_stages, p_set_name


os.environ["CUDA_VISIBLE_DEVICES"] = "0"

if __name__ == '__main__':
    # load data
    data_folder_path = './data-new/'

    center_name = 'SHANXI'  # DC--SHANXI EVC1--FUDAN  EVC2--YUNNAN

    all_p_name, all_p_patches, all_p_patch_LDs, all_p_patch_SDs, all_p_patch_RDs, all_p_patch_adc, all_LN_meta_labels, all_p_N_stages, p_set_names \
        = load_LN_data(center_name, data_folder_path)

    # create a new excel table for storing the preds and patient information
    wb = openpyxl.Workbook()  # 创建一个工作表
    ws = wb.active  # ws操作sheet页
    all_sheet = wb.create_sheet('1', 0)
    # write the table head (name of col)
    all_sheet.cell(1, 1, 'name')
    all_sheet.cell(1, 2, 'patch_pred_max')
    all_sheet.cell(1, 3, 'meta_status')
    all_sheet.cell(1, 4, 'N_stage')
    all_sheet.cell(1, 5, 'dataset')

    # ---------------------------------------------------------

    # load model to GPU
    device = torch.device("cuda:0")  # if torch.cuda.is_available() else "cpu"
    model_path = '/xiav/LN_diagnosis_pytorch/results-new-ban/Fusion_BCELoss_externalTest_date_240702/lr_0.001_bs_64_v0/model_auc_0.773_0.781_0.812_0.822_epoch_89.pkl'
    with open(model_path, 'rb') as f:
        model_dict = torch.load(f)
    model = ResNetSelf_Combine_ImgFeas_TwoOuts(1)
    model.load_state_dict(model_dict)
    # print(model)
    model = model.to(device)
    torch.set_grad_enabled(False)
    model.eval()

    all_patch_ind = 1

    for p_id in range(0, len(all_p_patches)):
        p_patches = all_p_patches[p_id]
        p_name = all_p_name[p_id]
        p_patch_SDs = all_p_patch_SDs[p_id]
        p_patch_LDs = all_p_patch_LDs[p_id]
        p_patch_RDs = all_p_patch_RDs[p_id]
        p_patch_adcs = all_p_patch_adc[p_id]
        LN_meta_labels = all_LN_meta_labels[p_id]
        p_N_stage = all_p_N_stages[p_id]
        p_set_name = p_set_names[p_id]
        p_preds = []
        all_patch_ind = all_patch_ind + 1
        for patch_id in range(0, len(p_patches)):
            p_patch = p_patches[patch_id]
            p_patch = T.ToTensor()(p_patch).unsqueeze(0)
            p_patch = torch.as_tensor(p_patch, dtype=torch.float)
            p_patch = p_patch.to(device)

            p_patch_LD = p_patch_LDs[patch_id]
            p_patch_LD = torch.as_tensor(p_patch_LD, dtype=torch.float)
            p_patch_LD = p_patch_LD.unsqueeze(0)
            p_patch_LD = p_patch_LD.to(device)

            p_patch_SD = p_patch_SDs[patch_id]
            p_patch_SD = torch.as_tensor(p_patch_SD, dtype=torch.float)
            p_patch_SD = p_patch_SD.unsqueeze(0)
            p_patch_SD = p_patch_SD.to(device)

            p_patch_RD = p_patch_RDs[patch_id]
            p_patch_RD = torch.as_tensor(p_patch_RD, dtype=torch.float)
            p_patch_RD = p_patch_RD.unsqueeze(0)
            p_patch_RD = p_patch_RD.to(device)

            p_patch_adc = p_patch_adcs[patch_id]
            p_patch_adc = torch.as_tensor(p_patch_adc, dtype=torch.float)
            p_patch_adc = p_patch_adc.unsqueeze(0)
            p_patch_adc = p_patch_adc.to(device)

            _, patch_combined_pred = model(p_patch, p_patch_LD, p_patch_SD, p_patch_RD, p_patch_adc)
            patch_combined_pred = patch_combined_pred.squeeze(0)
            patch_combined_pred = patch_combined_pred.cpu()
            patch_combined_pred = patch_combined_pred.numpy()
            patch_combined_pred = patch_combined_pred[0]
            p_preds.append(patch_combined_pred)

        all_sheet.cell(all_patch_ind, 1, p_name)
        all_sheet.cell(all_patch_ind, 2, max(p_preds))
        all_sheet.cell(all_patch_ind, 3, LN_meta_labels)
        all_sheet.cell(all_patch_ind, 4, p_N_stage)
        all_sheet.cell(all_patch_ind, 5, p_set_name)
        # -----------------------------------------------
        print('p_name', p_name, 'preds complete.')
    wb.save('./' + center_name + '_patient_preds.xlsx')