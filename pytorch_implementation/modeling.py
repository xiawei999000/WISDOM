# Author: xiaw@sibet.ac.cn
from __future__ import print_function, division
import torch
import torch.nn as nn
import torch.optim as optim
from torch.optim import lr_scheduler
import numpy as np
import torchvision
from torchvision import transforms as T
from torchvision import datasets, models
import matplotlib.pyplot as plt
import time
import math
import os
import copy
import random
import torch.utils.data as data
import pickle
from sklearn.metrics import roc_auc_score, f1_score, accuracy_score, recall_score, precision_score, roc_curve, confusion_matrix
from _collections import OrderedDict
import openpyxl
from sklearn.model_selection import KFold, StratifiedKFold
# import pandas
# from torch.utils.tensorboard import SummaryWriter
from Self_defined_Resnet_model_1 import ResNetSelf_Combine_ImgFeas_TwoOuts # test for smaller model performance
import monai
from monai.utils import set_determinism

def log_create():
    # INFO
    # data sheet
    info = openpyxl.Workbook()  # 创建一个工作表
    # ws = info.active  # ws操作sheet页
    all_sheet = info.create_sheet('1', 0)
    all_sheet.cell(1, 1, 'learning_rate')
    all_sheet.cell(1, 2, 'batch_size')
    all_sheet.cell(1, 3, 'epoch_id')
    all_sheet.cell(1, 4, 'loss_train')
    all_sheet.cell(1, 5, 'AUC_IVC')
    all_sheet.cell(1, 6, 'AUC_EVC1')
    all_sheet.cell(1, 7, 'AUC_EVC2')
    return info

class LN_Dataset(data.Dataset):
    # p_patch_size_feas_list : patches size features including short, long axis and ratio
    # p_Label_list: patient had LN meta (1) or not (0)
    def __init__(self, p_T2_patch_list, p_patch_LD_list, p_patch_SD_list, p_patch_RD_list, p_patch_adc_list, p_LN_meta_label_list, p_LN_meta_ratio):
        self.list_p_patch = p_T2_patch_list
        self.list_p_patch_LD = p_patch_LD_list
        self.list_p_patch_SD = p_patch_SD_list
        self.list_p_patch_RD = p_patch_RD_list
        self.list_p_patch_adc = p_patch_adc_list
        self.list_p_LN_meta_label = p_LN_meta_label_list
        self.list_p_LN_meta_ratio = p_LN_meta_ratio

    def __getitem__(self, idx):
        p_patch = self.list_p_patch[idx]
        p_patch_LD = self.list_p_patch_LD[idx]
        p_patch_SD = self.list_p_patch_SD[idx]
        p_patch_RD = self.list_p_patch_RD[idx]
        p_patch_adc = self.list_p_patch_adc[idx]
        p_LN_meta_label = self.list_p_LN_meta_label[idx]
        p_LN_meta_ratio = self.list_p_LN_meta_ratio[idx]
        return p_patch, p_patch_LD, p_patch_SD, p_patch_RD, p_patch_adc, p_LN_meta_label, p_LN_meta_ratio

    def __len__(self):
        return len(self.list_p_LN_meta_label)

def my_collate(batch):
    p_patches = []
    p_patch_LDs = []
    p_patch_SDs = []
    p_patch_RDs = []
    p_patch_adcs = []
    p_LN_meta_labels = []
    p_LN_meta_ratios = []
    for p_patch, p_patch_LD, p_patch_SD, p_patch_RD, p_patch_adc, p_LN_meta_label, p_LN_meta_ratio in batch:
        p_patches.append(p_patch)
        p_patch_LDs.append(p_patch_LD)
        p_patch_SDs.append(p_patch_SD)
        p_patch_RDs.append(p_patch_RD)
        p_patch_adcs.append(p_patch_adc)
        p_LN_meta_labels.append(p_LN_meta_label)
        p_LN_meta_ratios.append(p_LN_meta_ratio)
    return p_patches, p_patch_LDs, p_patch_SDs, p_patch_RDs, p_patch_adcs, p_LN_meta_labels, p_LN_meta_ratios

# ensure the reproducbility
def setup_seed(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    np.random.seed(seed)
    random.seed(seed)
    torch.backends.cudnn.deterministic = True

# load data
def load_LN_data(center_name, data_folder_path):
    # load original data
    all_patches = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'T2_2D_patches_correct_3channel.bin', "rb"))
    all_patch_LDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_long_diameter.bin', "rb"))
    all_patch_SDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_short_diameter.bin', "rb"))
    all_patch_RDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_ratio_diameter.bin', "rb"))
    all_patch_adc = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_adc_value.bin', "rb"))
    all_LN_meta_labels = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_label.bin', "rb"))
    all_LN_meta_ratios = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_LN_meta_ratio.bin', "rb"))

    # # exclude the bad data
    ban_list = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_ban_list.bin', "rb"))
    # ban_list = []
    all_p_ind = list(range(len(all_patches)))
    all_p_include_ind = []
    for ind in all_p_ind:
        if ind not in ban_list:
            all_p_include_ind.append(ind)

    all_patches = [all_patches[i] for i in all_p_include_ind]
    all_patch_LDs = [all_patch_LDs[i] for i in all_p_include_ind]
    all_patch_SDs = [all_patch_SDs[i] for i in all_p_include_ind]
    all_patch_RDs = [all_patch_RDs[i] for i in all_p_include_ind]
    all_patch_adc = [all_patch_adc[i] for i in all_p_include_ind]
    all_LN_meta_labels = [all_LN_meta_labels[i] for i in all_p_include_ind]
    all_LN_meta_ratios = [all_LN_meta_ratios[i] for i in all_p_include_ind]

    return all_patches, all_patch_LDs, all_patch_SDs, all_patch_RDs, all_patch_adc, all_LN_meta_labels, all_LN_meta_ratios #  ,

def load_LN_data_by_set(training_data_center_name, data_folder_path, set_name):
    # load original data for model building
    all_p_patches = pickle.load(open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/'
                                     + 'T2_2D_patches_correct_3channel.bin', "rb"))
    all_p_patch_LDs = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/Patches_long_diameter.bin', "rb"))
    all_p_patch_SDs = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/Patches_short_diameter.bin', "rb"))
    all_p_patch_RDs = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/Patches_ratio_diameter.bin', "rb"))
    all_p_patch_adc = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/Patches_adc_value.bin', "rb"))
    all_LN_meta_labels = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/P_label.bin', "rb"))
    all_LN_meta_ratios = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/P_LN_meta_ratio.bin', "rb"))

    set_ind = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/' + set_name + '_ind.bin',
             "rb"))

    # exclude the patients in ban list
    set_ind_selected = []
    ban_list = pickle.load(open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/P_ban_list.bin', "rb"))
    for ind in set_ind:
        if ind not in ban_list:
            set_ind_selected.append(ind)

    # arrange the training or validation set
    p_patches = [all_p_patches[i] for i in set_ind_selected]
    p_patch_LDs = [all_p_patch_LDs[i] for i in set_ind_selected]
    p_patch_SDs = [all_p_patch_SDs[i] for i in set_ind_selected]
    p_patch_RDs = [all_p_patch_RDs[i] for i in set_ind_selected]
    p_patch_adc = [all_p_patch_adc[i] for i in set_ind_selected]
    p_LN_meta_label = [all_LN_meta_labels[i] for i in set_ind_selected]
    p_LN_meta_ratio = [all_LN_meta_ratios[i] for i in set_ind_selected]

    return p_patches, p_patch_LDs, p_patch_SDs, p_patch_RDs, p_patch_adc, p_LN_meta_label, p_LN_meta_ratio


def evaluate(y_true, y_pred, digits=3, cutoff='auto'):
    '''
    calculate several metrics of predictions
    Args:
        y_true: list, labels
        y_pred: list, predictions
        digits: The number of decimals to use when rounding the number. Default is 4
        cutoff: float or 'auto'

    Returns:
        evaluation: dict
    '''

    if cutoff == 'auto':
        fpr, tpr, thresholds = roc_curve(y_true, y_pred)
        youden = tpr-fpr
        cutoff = thresholds[np.argmax(youden)]

    y_pred_t = [1 if i > cutoff else 0 for i in y_pred]

    evaluation = OrderedDict()
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred_t).ravel()
    evaluation['AUC'] = round(roc_auc_score(y_true, y_pred), digits)
    evaluation['ACC'] = round(accuracy_score(y_true, y_pred_t), digits)
    # evaluation['recall'] = round(recall_score(y_true, y_pred_t), digits)
    evaluation['SEN'] = round(tp / (tp + fn), digits)
    evaluation['SPE'] = round(tn / (tn + fp), digits)
    # evaluation['F1'] = round(f1_score(y_true, y_pred_t), digits)
    evaluation['cutoff'] = cutoff

    return evaluation


def train_model(device, dataloaders, model, criterions, optimizer, scheduler,
                batch_size, learning_rate, num_epochs, work_dir, log_path, log_modelling, index_record):

    since = time.time()
    # online data augmentation in training
    random_transform = T.Compose([
        # 1 Horizontal Flip
        T.RandomHorizontalFlip(),
        # 2 Vertical Flip
        T.RandomVerticalFlip(),
        # 3 RandomRotation
        T.RandomRotation(10),
        ## 4 brightness adjust
        # T.ColorJitter(brightness=0.2, contrast=0, saturation=0, hue=0)
    ])

    # folder to store models, logs
    # folder named by hyper-paramters
    results_save_folder = work_dir + 'lr_{}_bs_{}/'.format(learning_rate, batch_size)
    if not os.path.exists(results_save_folder):
        os.makedirs(results_save_folder)

    # the sheet to store infomation
    shenames = log_modelling.get_sheet_names()
    all_sheet = log_modelling[shenames[0]]
    #
    AUC_train = 0.0
    AUC_val = 0.0
    AUC_EVC1 = 0.0
    AUC_EVC2 = 0.0

    best_val_auc = 0.0

    for epoch in range(num_epochs):
        epoch_id = epoch + 1
        print('Epoch {}/{}'.format(epoch_id, num_epochs))
        print('-' * 10)
        # optimizer.print_lr()
        print('running learning rate: ', optimizer.param_groups[0]['lr'])
        # runnning_learning_rate = optimizer.state_dict()['param_groups'][0]['lr']
        # print('running learning rate: ', runnning_learning_rate)
        epoch_start_time = time.time()
        # Each epoch has a training and validation and test phase
        # store the performance index for each model
        # write log
        all_sheet.cell(index_record, 1, learning_rate)  # 'learning_rate')
        all_sheet.cell(index_record, 2, batch_size)  # 'batch_size')
        all_sheet.cell(index_record, 3, epoch_id)  # 'epoch_id')

        for phase in ['train', 'val', 'EVC1', 'EVC2']:
            # phase_flag = 0
            if phase == 'train':
                model.train()  # Set model to training mode
                phase_flag = 0
            elif phase == 'val':
                model.eval()   # Set model to evaluate mode
                phase_flag = 1
            elif phase == 'EVC1':
                model.eval()   # Set model to evaluate mode
                phase_flag = 2
            elif phase == 'EVC2':
                model.eval()   # Set model to evaluate mode
                phase_flag = 3
            # else:
            #     model.eval()  # Set model to evaluate mode
            #     phase_flag = 2
            torch.set_grad_enabled(phase == 'train')

            size = len(dataloaders[phase_flag].dataset)
            num_mini_batches = math.ceil(size / batch_size)
            print('num_mini_batches = ', num_mini_batches)

            running_loss = 0.0

            epoch_outputs_p = []
            epoch_labels_p = []

            # Iterate over data.
            for batch, (p_patches_batch, p_patch_LDs_batch, p_patch_SDs_batch, p_patch_RDs_batch, p_patch_adc_batch, p_LN_meta_label_batch, p_LN_meta_ratio_batch) \
                    in enumerate(dataloaders[phase_flag]):

                # two labels
                LN_meta_label_p_batch = torch.tensor(p_LN_meta_label_batch, dtype=torch.float)
                LN_meta_label_p_batch = LN_meta_label_p_batch.to(device)

                LN_meta_ratio_p_batch = torch.tensor(p_LN_meta_ratio_batch, dtype=torch.float)
                LN_meta_ratio_p_batch = LN_meta_ratio_p_batch.to(device)

                # store model outputs for a batch of patients in img level
                img_max_outputs_p_batch = []
                img_max_outputs_p_batch = torch.tensor(img_max_outputs_p_batch)
                img_max_outputs_p_batch = img_max_outputs_p_batch.to(device)

                img_avg_outputs_p_batch = []
                img_avg_outputs_p_batch = torch.tensor(img_avg_outputs_p_batch)
                img_avg_outputs_p_batch = img_avg_outputs_p_batch.to(device)

                # store model outputs for a batch of patients in combined features level
                combined_max_outputs_p_batch = []
                combined_max_outputs_p_batch = torch.tensor(combined_max_outputs_p_batch)
                combined_max_outputs_p_batch = combined_max_outputs_p_batch.to(device)

                combined_avg_outputs_p_batch = []
                combined_avg_outputs_p_batch = torch.tensor(combined_avg_outputs_p_batch)
                combined_avg_outputs_p_batch = combined_avg_outputs_p_batch.to(device)

                # load per patient's patches
                for p_id in range(0, len(p_patches_batch)):
                    p_patches = p_patches_batch[p_id]
                    p_patch_LDs = p_patch_LDs_batch[p_id]
                    p_patch_SDs = p_patch_SDs_batch[p_id]
                    p_patch_RDs = p_patch_RDs_batch[p_id]
                    p_patch_adcs = p_patch_adc_batch[p_id]
                    p_LN_meta_label = p_LN_meta_label_batch[p_id]
                    p_LN_meta_ratio = p_LN_meta_ratio_batch[p_id]

                    # forward
                    # generate the patch predictions for each patient in img level
                    img_outputs_p_patches = []
                    img_outputs_p_patches = torch.tensor(img_outputs_p_patches)
                    img_outputs_p_patches = img_outputs_p_patches.to(device)

                    # generate the patch predictions for each patient in combined features level
                    combined_outputs_p_patches = []
                    combined_outputs_p_patches = torch.tensor(combined_outputs_p_patches)
                    combined_outputs_p_patches = combined_outputs_p_patches.to(device)

                    patch_num = len(p_patches)
                    for patch_id in range(0, patch_num):
                        # extract each LN patch and input into model
                        p_patch = p_patches[patch_id]
                        p_patch = T.ToTensor()(p_patch)
                        if phase == 'train': #  or 'val' # online data augmentation in training phase
                            p_patch = random_transform(p_patch)
                        p_patch = p_patch.unsqueeze(0)
                        p_patch = torch.as_tensor(p_patch, dtype=torch.float)
                        p_patch = p_patch.to(device)

                        # extract each LN patch features and input into model
                        p_patch_LD = p_patch_LDs[patch_id]/10.0
                        p_patch_LD = torch.as_tensor(p_patch_LD, dtype=torch.float)
                        p_patch_LD = p_patch_LD.unsqueeze(0)
                        p_patch_LD = p_patch_LD.to(device)

                        p_patch_SD = p_patch_SDs[patch_id]/10.0
                        p_patch_SD = torch.as_tensor(p_patch_SD, dtype=torch.float)
                        p_patch_SD = p_patch_SD.unsqueeze(0)
                        p_patch_SD = p_patch_SD.to(device)

                        p_patch_RD = p_patch_RDs[patch_id]
                        p_patch_RD = torch.as_tensor(p_patch_RD, dtype=torch.float)
                        p_patch_RD = p_patch_RD.unsqueeze(0)
                        p_patch_RD = p_patch_RD.to(device)

                        p_patch_adc = p_patch_adcs[patch_id]/100.0
                        p_patch_adc = torch.as_tensor(p_patch_adc, dtype=torch.float)
                        p_patch_adc = p_patch_adc.unsqueeze(0)
                        p_patch_adc = p_patch_adc.to(device)

                        [patch_img_pred, patch_combined_pred] = model(p_patch, p_patch_LD, p_patch_SD, p_patch_RD, p_patch_adc)
                        combined_outputs_p_patches = torch.cat((combined_outputs_p_patches, patch_combined_pred))

                    ####  get the max prediction of all patches as per patient prediction  ####
                    combined_outputs_p_patch_max, _ = torch.max(combined_outputs_p_patches, 0)
                    combined_outputs_p_patch_avg = torch.mean(combined_outputs_p_patches, 0)

                    combined_max_outputs_p_batch = torch.cat(
                        (combined_max_outputs_p_batch, combined_outputs_p_patch_max))
                    combined_avg_outputs_p_batch = torch.cat(
                        (combined_avg_outputs_p_batch, combined_outputs_p_patch_avg))

                loss_combined_max = criterions['loss_combined_max'](combined_max_outputs_p_batch, LN_meta_label_p_batch) # LOSS for max pred in combined feature level
                loss_combined_avg = criterions['loss_combined_avg'](combined_avg_outputs_p_batch, LN_meta_ratio_p_batch) # LOSS for avg pred in combined feature level
                loss = loss_combined_max + loss_combined_avg  # Twolables+FinalFusion

                if phase == 'train':
                    optimizer.zero_grad()
                    loss.backward()
                    optimizer.step()
                # print internal information
                loss_item, current = loss.item(), batch * batch_size
                print(f"batch {batch+1} - loss: {loss_item:>4f}  [{current:>5d}/{size:>5d}]")

                running_loss += loss.item()  # total loss for one epoch

                epoch_outputs_p.extend(combined_max_outputs_p_batch.tolist())
                epoch_labels_p.extend(p_LN_meta_label_batch)

            # statistics of one epoch
            epoch_loss = running_loss / num_mini_batches

            if phase == 'train':
                epoch_evaluations = evaluate(np.array(epoch_labels_p), np.array(epoch_outputs_p), digits=3)
                epoch_AUC = epoch_evaluations["AUC"]
                epoch_ACC = epoch_evaluations["ACC"]
                epoch_SEN = epoch_evaluations["SEN"]
                epoch_SPE = epoch_evaluations["SPE"]
                print('{} Loss: {:.4f} {} AUC:{:.4f} {} ACC: {:.4f} {} SEN: {:.4f} {} SPE: {:.4f} {} '.format(
                    phase, epoch_loss, phase, epoch_AUC, phase, epoch_ACC, phase, epoch_SEN, phase, epoch_SPE, phase))
                all_sheet.cell(index_record, 4, epoch_loss)
                AUC_train = epoch_AUC

            if phase == 'val':
                epoch_evaluations = evaluate(np.array(epoch_labels_p), np.array(epoch_outputs_p), digits=3)
                epoch_AUC = epoch_evaluations["AUC"]
                epoch_ACC = epoch_evaluations["ACC"]
                epoch_SEN = epoch_evaluations["SEN"]
                epoch_SPE = epoch_evaluations["SPE"]
                print('{} Loss: {:.4f} {} AUC:{:.4f} {} ACC: {:.4f} {} SEN: {:.4f} {} SPE: {:.4f}'.format(
                    phase, epoch_loss, phase, epoch_AUC, phase, epoch_ACC, phase, epoch_SEN, phase, epoch_SPE))
                all_sheet.cell(index_record, 5, epoch_AUC)
                AUC_val = epoch_AUC

            if phase == 'EVC1':
                epoch_evaluations = evaluate(np.array(epoch_labels_p), np.array(epoch_outputs_p), digits=3)
                epoch_AUC = epoch_evaluations["AUC"]
                epoch_ACC = epoch_evaluations["ACC"]
                epoch_SEN = epoch_evaluations["SEN"]
                epoch_SPE = epoch_evaluations["SPE"]
                print('{} Loss: {:.4f} {} AUC:{:.4f} {} ACC: {:.4f} {} SEN: {:.4f} {} SPE: {:.4f}'.format(
                    phase, epoch_loss, phase, epoch_AUC, phase, epoch_ACC, phase, epoch_SEN, phase, epoch_SPE))
                all_sheet.cell(index_record, 6, epoch_AUC)
                AUC_EVC1 = epoch_AUC

            if phase == 'EVC2':
                epoch_evaluations = evaluate(np.array(epoch_labels_p), np.array(epoch_outputs_p), digits=3)
                epoch_AUC = epoch_evaluations["AUC"]
                epoch_ACC = epoch_evaluations["ACC"]
                epoch_SEN = epoch_evaluations["SEN"]
                epoch_SPE = epoch_evaluations["SPE"]
                print('{} Loss: {:.4f} {} AUC:{:.4f} {} ACC: {:.4f} {} SEN: {:.4f} {} SPE: {:.4f}'.format(
                    phase, epoch_loss, phase, epoch_AUC, phase, epoch_ACC, phase, epoch_SEN, phase, epoch_SPE))
                all_sheet.cell(index_record, 7, epoch_AUC)
                AUC_EVC2 = epoch_AUC

            # update learning rate after calculating epoch index in validation data
            if phase == 'val':
                # scheduler.step(epoch_loss)  # for the scheduler required loss of epoch
                scheduler.step()

        # save each model
        # record the train, val and test auc, batch_size, lr, epoch
        torch.save(model.state_dict(),
                    results_save_folder +
                       'model_auc_{:.3f}_{:.3f}_{:.3f}_{:.3f}_epoch_{}.pkl'.format(AUC_train, AUC_val, AUC_EVC1, AUC_EVC2, epoch+1))

        # write log for the next epoch
        index_record = index_record + 1

        # computation time of each epoch
        epoch_end_time = time.time()
        print('epoch computation time: ', str(epoch_end_time-epoch_start_time))
    # save the training log
    log_modelling.save(filename=log_path)
    # computation time for entire training
    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('Best val AUC: {:.4f}'.format(best_val_auc))
    print()
    print()
    #
    # # load best model weights
    # model.load_state_dict(best_model_wts)
    #
    # return model

if __name__ == '__main__':
    # ensure the reproducbility
    my_seed = 702
    set_determinism(seed=my_seed)
    # setup_seed(my_seed)

    date = "0702"
    # data folder
    data_folder_path = './data-new/'
    result_folder_path = './results-new-ban/'  # using the data corrected

    ######################## load training, validation and test data  ########################
    # load original data for model building
    center1_name = 'SHANXI'
    center2_name = 'FUDAN'
    center3_name = 'ZFY'

    # load data
    p_patches_train, p_patch_LDs_train, p_patch_SDs_train, p_patch_RDs_train, p_patch_adc_train, p_LN_meta_label_train, p_LN_meta_ratio_train \
        = load_LN_data_by_set(center1_name, data_folder_path, 'training')
    p_patches_val, p_patch_LDs_val, p_patch_SDs_val, p_patch_RDs_val, p_patch_adc_val, p_LN_meta_label_val, p_LN_meta_ratio_val \
        = load_LN_data_by_set(center1_name, data_folder_path, 'val')
    p_patches_EVC1, p_patch_LDs_EVC1, p_patch_SDs_EVC1, p_patch_RDs_EVC1, p_patch_adc_EVC1, p_LN_meta_label_EVC1, p_LN_meta_ratio_EVC1 \
        = load_LN_data(center2_name, data_folder_path)
    p_patches_EVC2, p_patch_LDs_EVC2, p_patch_SDs_EVC2, p_patch_RDs_EVC2, p_patch_adc_EVC2, p_LN_meta_label_EVC2, p_LN_meta_ratio_EVC2 \
        = load_LN_data(center3_name, data_folder_path)

    ################# model training ########################
    # initialize the GPU
    device = torch.device("cuda:0") #  if torch.cuda.is_available() else "cpu"
    # set hyper-parameters for model training
    batch_size_list = [32, 16]
    learning_rate_list = [1e-2, 1e-3] #[1e-2, 4e-3, 1e-3, 4e-4]
    epoch = 200

    # define loss function
    loss_fuc = 'BCELoss'
    criterions = OrderedDict()
    criterions['loss_combined_max'] = nn.BCELoss()  # LOSS for max pred in combined feature level
    criterions['loss_combined_avg'] = nn.MSELoss()  # LOSS for avg pred in combined feature level

    # the model in use
    model_name = 'Fusion'

    # path to store results
    work_dir = result_folder_path + model_name + '_' + loss_fuc + '_externalTest' + '_date_' + date + '/'
    if not os.path.exists(work_dir):
        os.makedirs(work_dir)
    log_path = work_dir + '/log.xlsx'
    log_modelling = log_create()
    log_modelling.save(filename=log_path)
    index_record = 2

    # batch size and learning rate as hyper-parameters
    for batch_size in batch_size_list:
        for learning_rate in learning_rate_list:
            print(
                "Number of samples in train, validation, EVC1, and EVC2 are %d, %d, %d and %d."
                % (len(p_patches_train), len(p_patches_val), len(p_patches_EVC1), len(p_patches_EVC2)))

            # load pytorch dataset
            train_dataset = LN_Dataset(p_T2_patch_list=p_patches_train,
                                       p_patch_LD_list=p_patch_LDs_train,
                                       p_patch_SD_list=p_patch_SDs_train,
                                       p_patch_RD_list=p_patch_RDs_train,
                                       p_patch_adc_list=p_patch_adc_train,
                                       p_LN_meta_label_list=p_LN_meta_label_train,
                                       p_LN_meta_ratio=p_LN_meta_ratio_train)
            train_dataloader = torch.utils.data.DataLoader(train_dataset, batch_size=batch_size, shuffle=True,
                                                           num_workers=4, pin_memory=True, collate_fn=my_collate)

            val_dataset = LN_Dataset(p_T2_patch_list=p_patches_val,
                                     p_patch_LD_list=p_patch_LDs_val,
                                     p_patch_SD_list=p_patch_SDs_val,
                                     p_patch_RD_list=p_patch_RDs_val,
                                     p_patch_adc_list=p_patch_adc_val,
                                     p_LN_meta_label_list=p_LN_meta_label_val,
                                     p_LN_meta_ratio=p_LN_meta_ratio_val)
            val_dataloader = torch.utils.data.DataLoader(val_dataset, batch_size=batch_size,
                                                           num_workers=4, pin_memory=True, collate_fn=my_collate)

            EVC1_dataset = LN_Dataset(p_T2_patch_list=p_patches_EVC1,
                                     p_patch_LD_list=p_patch_LDs_EVC1,
                                     p_patch_SD_list=p_patch_SDs_EVC1,
                                     p_patch_RD_list=p_patch_RDs_EVC1,
                                     p_patch_adc_list=p_patch_adc_EVC1,
                                     p_LN_meta_label_list=p_LN_meta_label_EVC1,
                                     p_LN_meta_ratio=p_LN_meta_ratio_EVC1)
            EVC1_dataloader = torch.utils.data.DataLoader(EVC1_dataset, batch_size=batch_size,
                                                           num_workers=4, pin_memory=True, collate_fn=my_collate)


            EVC2_dataset = LN_Dataset(p_T2_patch_list=p_patches_EVC2,
                                     p_patch_LD_list=p_patch_LDs_EVC2,
                                     p_patch_SD_list=p_patch_SDs_EVC2,
                                     p_patch_RD_list=p_patch_RDs_EVC2,
                                     p_patch_adc_list=p_patch_adc_EVC2,
                                     p_LN_meta_label_list=p_LN_meta_label_EVC2,
                                     p_LN_meta_ratio=p_LN_meta_ratio_EVC2)
            EVC2_dataloader = torch.utils.data.DataLoader(EVC2_dataset, batch_size=batch_size,
                                                           num_workers=4, pin_memory=True, collate_fn=my_collate)



            dataloaders = [train_dataloader, val_dataloader, EVC1_dataloader, EVC2_dataloader]

            # load model to GPU
            model_ft = ResNetSelf_Combine_ImgFeas_TwoOuts(1)
            # print(model_ft)
            model_ft = model_ft.to(device)

            # Observe that all parameters are being optimized
            optimizer_ft = optim.AdamW(model_ft.parameters(), lr=learning_rate)
            # lr_scheduler
            # Decay LR by a factor of 0.5 every 5 epochs
            # exp_lr_scheduler = lr_scheduler.StepLR(optimizer_ft, step_size=5, gamma=0.5)
            exp_lr_scheduler = lr_scheduler.ReduceLROnPlateau(optimizer_ft, mode='min', factor=0.5, patience=10, threshold=0.001,
                                                       threshold_mode='abs', cooldown=5, min_lr=1e-7, eps=1e-08, verbose=False)

            CosineAnnealingLR_scheduler = lr_scheduler.CosineAnnealingLR(optimizer_ft, T_max=(epoch * len(
                p_LN_meta_label_train)) // batch_size / 10, eta_min=1e-8, verbose=True)


            torch.backends.cudnn.benchmark = True
            train_model(device, dataloaders, model_ft, criterions,
                        optimizer_ft, CosineAnnealingLR_scheduler, batch_size,
                        learning_rate, epoch, work_dir, log_path, log_modelling, index_record)
            index_record = index_record + epoch

